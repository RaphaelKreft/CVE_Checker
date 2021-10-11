import os
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

"""
This file contains methods to read and write from/to csv and excel files. they dump and try to read a so
called software-dictionary which contains the following fields: key=swName, value=date_of_last_check
"""


def table_to_array(table):
    """
    Takes a table as pyxl-object and coverts it into an 2D array.
    """
    array = []
    for row in table:
        array.append([cell.value for cell in row])
    return array


def read_software_data(input_file: str, name_col: int, date_col: int, delimiter: str = "", excel=False):
    """
    This function reads the software-data from a file.
    """
    if not os.path.isfile(input_file):
        raise IOError(f"File {input_file} doesn't exist!")
    else:
        software_data = []
        with open(input_file, mode="r", encoding="utf-8") as f:
            if excel:
                # if we read from excel file
                workbook = load_workbook(input_file, data_only=True)
                ws = workbook.active
                table = list(ws.tables.values())[0]  # here we assume that just one table in the right format exists
                return table_to_array(ws[table.ref])
            else:
                # if we read from csv file
                csv_reader = csv.reader(f, delimiter=delimiter)
                for row in csv_reader:
                    software_data.append([row[name_col], row[date_col] if len(row) >= 2 else ""])
                return software_data


def prettify_sheet(sheet, datecol_num):
    """
    This Method should format a sheet(in our case the input data sheet) in a specific way. It makes the Date column wider so that one can read
    the date.
    """
    # here relation of Date_column, Found_Items, Highest Severity to their width is stored
    num_dim_relations = [(datecol_num, 20), (2, 20), (3, 15)]
    for col_num, width in num_dim_relations:
        sheet.column_dimensions[get_column_letter(col_num + 1)].width = width


def dumb_updated_data(output_file, software_list: list, delimiter: str = ",", excel=False, date_col=1):
    """
    This method writes the information regarding all software back into a file.
    """
    # delete existing file
    if os.path.isfile(output_file):
        os.remove(output_file)
    # create new file and write to it
    if excel:
        wb = Workbook()
        sheet = wb.active
        # create raw data
        # The first row of a table is always the table headers, these are kept!
        for row in software_list:
            sheet.append(row)
        # make table out of it
        tab = Table(displayName="CVE_Info_Table", ref=f"A1:D{len(software_list)}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)
        prettify_sheet(sheet, datecol_num=date_col)
        wb.save(output_file)
    else:
        with open(output_file, mode='w') as f:
            csv_writer = csv.writer(f, delimiter=delimiter)
            csv_writer.writerow(["Name", "Last check", "New Items", "Highest severity"])
            for row in software_list:
                csv_writer.writerow(row)


if __name__ == "__main__":
    # test read of data from excel sheet
    data = table_to_array(read_software_data("examples/test.xlsx", 0, 1, excel=True))
    print(data)
    # test dump
    dumb_updated_data("examples/output.xlsx", data, excel=True)
